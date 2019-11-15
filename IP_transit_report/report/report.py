#!/usr/bin/python
# -*- coding: UTF-8 -*-

import calendar
import configparser
import csv
import datetime
import json
import sys
import pyinstaller

import time
from time import strftime, localtime
from urllib.request import urlopen

import openpyxl


# 打印当前时间
def getNowTime():
    nowTime = strftime("%Y-%m-%d %H:%M:%S", localtime())
    return nowTime


'''
获取临时的结果,结果从大到小排序
type：In 获取in流量，Out获取out流量
'''


def getListByInOrOut(type="In", device="", interface="", ):
    tempList = []
    try:
        if (type == "In"):
            url = param.InOctets2bps
        else:
            url = param.OutOctets2bps
        startTime = int(time.mktime(datetime.datetime.strptime(param.startTime, '%Y%m%d').timetuple()))
        endTime = int(time.mktime(datetime.datetime.strptime(param.endTime, '%Y%m%d').timetuple()))
        stepFrequency = int((endTime - startTime) / 10009)
        url = url.replace("device", device)
        url = url.replace("interface", interface)
        url = url.replace("startTime", str(startTime))
        url = url.replace("endTime", str(endTime))
        url = url.replace("stepFrequency", str(stepFrequency))
        tempResult = get_result_page(url)
        values = json.loads(tempResult)['data']['result'][0]['values']
        for i in range(0, list(values).__len__()):
            temparr = [values[i][0], int(float(values[i][1]))]
            tempList.append(temparr)
        tempList = sorted(tempList, key=(lambda x: x[1]), reverse=True)
    except Exception as e:
        print("url:" + url)
        raise
    return tempList


'''
根据list结果获取95%最大值最小值 平均数
'''


def getAvgMinMaxByList(list):
    byList = {}
    nsum = 0
    for i in range(0, list.__len__()):
        nsum += list[i][1]
    byList["avg"] = round(nsum / list.__len__() / 1000 / 1024, 2)
    byList["min"] = round(list[round(list.__len__() * 0.95)][1] / 1000 / 1024, 2)
    byList["max"] = round(list[list.__len__() - round(list.__len__() * 0.95)][1] / 1000 / 1024, 2)
    return byList


'''
根据 接口和设备获取参数
'''


def getResult(bandwidth="", device="", interface=""):
    result = []
    inList = getListByInOrOut(type="In", device=device, interface=interface)
    outList = getListByInOrOut(type="Out", device=device, interface=interface)
    inListR = getAvgMinMaxByList(inList)
    outListR = getAvgMinMaxByList(outList)
    result.append(inListR["avg"])
    result.append(outListR["avg"])
    result.append(inListR["max"])
    result.append(outListR["max"])
    imax95 = round(float(inListR["max"]) * 100 / int(bandwidth), 2)
    omax95 = round(float(outListR["max"]) * 100 / int(bandwidth), 2)
    result.append(imax95)
    result.append(omax95)
    if (imax95 >= omax95):
        result.append(imax95)
    else:
        result.append(omax95)
    return result


# 读取execl 改变excel内容
def readExcel(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for i in range(1, ws.max_row + 1):
        bandwidth = ws.cell(row=i, column=3).value
        device = ws.cell(row=i, column=4).value
        interface = ws.cell(row=i, column=5).value
        if (bandwidth != None and device != None and interface != None and str(bandwidth).isdigit()):
            try:
                print("device:" + device + "\tinterface:" + interface)
                result = getResult(bandwidth=bandwidth, device=device, interface=interface)
            except Exception as e:
                # print(e)
                continue
            ws.cell(row=i, column=6, value=str(result[0]))
            ws.cell(row=i, column=7, value=str(result[1]))
            ws.cell(row=i, column=8, value=str(result[2]))
            ws.cell(row=i, column=9, value=str(result[3]))
            ws.cell(row=i, column=10, value=str(result[4]))
            ws.cell(row=i, column=11, value=str(result[5]))
            ws.cell(row=i, column=12, value=str(result[6]))
    wb.save(filename)
    wb.close()


# 访问
def get_result_page(url):
    res = urlopen(url).read()
    return str(res, 'utf-8')


## 加载conf里面的配置
def getConfig(configName='config.conf'):
    config = configparser.ConfigParser()
    config.read(configName, encoding='UTF-8-sig')
    param.startTime = str(config.get('time_config', 'startTime'))
    param.endTime = str(config.get('time_config', 'endTime'))
    param.excelFileName = str(config.get('excel_config', 'excelFileName'))
    param.InOctets2bps = str(config.get('default_config', 'InOctets2bps')[1:-1])
    param.OutOctets2bps = str(config.get('default_config', 'OutOctets2bps')[1:-1])


class confParam:
    startTime = ""
    endTime = ""
    excelFileName = ""
    InOctets2bps = ""
    OutOctets2bps = ""


'''
获取上一个周五的日期
'''


def getLastFriday(nowDate=""):
    oneday = datetime.timedelta(days=1)
    if (nowDate == ""):
        lastFriday = datetime.date.today()
    else:
        lastFriday = datetime.datetime.strptime(nowDate, '%Y%m%d')
        if (lastFriday.weekday() == calendar.FRIDAY):
            lastFriday -= oneday
    while lastFriday.weekday() != calendar.FRIDAY:
        lastFriday -= oneday
    return lastFriday.strftime('%Y%m%d')


class Logger(object):
    def __init__(self, filename='log.txt', stream=sys.stdout):
        self.terminal = stream
        self.log = open(filename, 'a', encoding='utf8')

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        pass


sys.stdout = Logger('log.txt', sys.stdout)
sys.stderr = Logger('log.txt', sys.stderr)

if __name__ == '__main__':
    param = confParam()
    getConfig()
    # print("excelFileName:"+param.excelFileName)
    # print("InOctets2bps:"+param.InOctets2bps)
    # print("OutOctets2bps:"+param.OutOctets2bps)
    if (param.endTime == ""):
        param.endTime = getLastFriday()
        param.startTime = getLastFriday(param.endTime)
    readExcel(param.excelFileName)
