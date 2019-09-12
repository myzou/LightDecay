#!/usr/bin/python
# -*- coding: UTF-8 -*-

import base64
import rsa
import time
from time import strftime, localtime
import openpyxl
from openpyxl import load_workbook
from urllib.request import urlopen
import sys
import json
import random
import pymysql
from datetime import datetime

executeLog = ""
deviceSize = 0

class Logger(object):
    def __init__(self, filename='console.log', stream=sys.stdout):
        self.terminal = stream
        self.log = open(filename, 'a',encoding='utf8')
    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
    def flush(self):
        pass
sys.stdout = Logger('console.log', sys.stdout)
sys.stderr = Logger('console.log', sys.stderr)


def get_result_page(url):
    res = urlopen(url).read()
    return str(res,'utf-8')


def encode_rsa(message, APIURL, ip, port):
    publicKey="-----BEGIN RSA PUBLIC KEY-----\nMIIBCgKCAQEAh/hChlpkRA+zkkEB8ZoCrVcsYsbFSXYoTBKlrdCu0LiGeKs2T+U7\nkyZ/WZ8GP498PIucz6GYN03BWOOPe5nHWfwO05XqTid6+0ni+Bfy4Ev0FyCOQsod\nmQpH4ytgn0UOp8BZyJTwdN4rtMcuY/FFnyAsFpg9+F0DtlM/dVMj/UcWaMiZIBaa\n35XkXoPa+ng8Z7ORVOPiRHXfMGrlb9gZWF5XHN1SHNBKha1uF3wexDHVm4+k3Hvb\naZFkHrgLndaYuCPPtetnSNUOw2iaiNo7Nfze1Y4ACyfvRczHEGxFEC9X7tj/Rcy2\ngmC9JCErxDQAHitX8DJrKtoeSJN8GvZZJQIDAQAB\n-----END RSA PUBLIC KEY-----\n"
    pubkey = rsa.PublicKey.load_pkcs1(publicKey)
    #公钥从文件中读取
    # with open(pub_file, 'r') as f:
    #     print(f.read().encode())
    #     pubkey = rsa.PublicKey.load_pkcs1(f.read().encode())

    aa = str(base64.b64encode(rsa.encrypt(message.encode(), pubkey)),'utf-8')
    # 登录GGW
    login = APIURL+'/GetLoginSession/RSA?crypto_sign=' + aa

    #登录PE并执行命令
    exec = APIURL+'/ExecuteCommand/RSA?crypto_sign='+aa+ \
        '&&command=%20show%20interfaces%20diagnostics%20optics%20{0}%20&&ip={1}'.format(port,ip)
    return login,exec


def get_message(username,password,sign):
    message='username=%s&&password=%s&&sign=%s&&timestamp=%s' % (username, password, sign, str(int(time.time())))
    return message

# 读取excel 执行了
def read_xls(filename):
    wb = load_workbook(filename)
    tempExecuteLog=""
    ws = wb.active
    username=ws.cell(row=2, column=7).value
    password=ws.cell(row=2, column=8).value
    APIURL=ws.cell(row=4, column=7).value
    global deviceSize
    deviceSize=ws.max_row-1
    if password==None or username==None :
        print("请在对应位置输入账号密码")
    if APIURL==None:
        print("APIURL地址不能为空：")
    if APIURL != None and password!=None and username!=None :
        for i in range(2, ws.max_row+1):
            if ws.cell(row=i, column=1).value==None:
                deviceSize-=1
                continue
            device_name = ws.cell(row=i, column=1).value
            ip = ws.cell(row=i, column=2).value
            port = ws.cell(row=i, column=3).value
            print(device_name,ip,port)
            mg = get_message(username, password, '123456')
            login, exec = encode_rsa(mg, APIURL, ip, port)
            print(login)
            print(exec)
            try:
                print(get_result_page(login))
                ss = get_result_page(exec)
                print(ss)
                returnData = json.loads(ss)
                tempExecuteLog+="=========" + device_name + "======"+ip+"===============\n"+device_name+">show interfaces diagnostics optics " + port + "\n"
                tempExecuteLog+= device_name+">"+returnData['data']+"\n"
            except Exception as e:
                tempExecuteLog+="异常："+str(e)+"\n";
                print(e)
                print('login or get information fail.')
                ws.cell(row=i, column=4, value='Login fail')
                wb.save(filename)
                print('---------------------------------------------------------------------------------')
                continue
            try:
                ws.cell(row=i, column=6, value=str(getNowTime()))
                if json.loads(ss)['data']=="" :
                    ws.cell(row=i, column=4, value='没有查询到内容')
                else:
                    if judge_json(ss) is True:
                        ws.cell(row=i,column=4,value='OK')
                    elif judge_json(ss) is False:
                        ws.cell(row=i,column=4,value='Not all off')
            except Exception as e:
                tempExecuteLog+="异常："+str(e)+"\n";
                print('get data exception. ')
                ws.cell(row=i,column=4,value='Exception')
            tempExecuteLog+="--------------------------------------------------------------------------------------------------\n"

            wb.save(filename)
            print('###################-save success-########################')
    global executeLog
    executeLog+="查询时间："+getNowTime()+"\n"
    executeLog+=tempExecuteLog


def judge_json(json_str):
    data = str(json.loads(json_str)['data']).split('\n')
    flg = True
    for i in range(6, 26):
        if data[i].split(':')[1].strip() != 'Off':
            flg = False
            break
    return flg

def random_sign(num):
    return "".join(random.sample('zyxwvutsrqponmlkjihgfedcba', num))


def test():
    print("this is test method")
    tempExecuteLog="这是一个测试方法"
    global executeLog
    executeLog += tempExecuteLog

#读取execl 改变excel内容
def readExcel(filename):
     data=queryLogin();
     wb = openpyxl.load_workbook(filename)
     ws = wb.active
     for i in range(2,ws.max_row+1):
          if ws.cell(row=i, column=1).value == None:
               continue
          device_name = ws.cell(row=i, column=1).value
          ip = ws.cell(row=i, column=2).value
          port = ws.cell(row=i, column=3).value
          print(i,device_name,ip,port)
          now_ip=getIp(data,device_name)
          ws.cell(row=i,column=2,value=now_ip)
          ws.cell(row=i, column=5,value=str(getNowTime()))
     wb.save(filename)
     wb.close()

##单个查询内容
def queryLogin():
     # 打开数据库连接
     db = pymysql.connect("218.97.9.147", "nss_script", "ZX4XdUpH", "nm_shared_info", charset='utf8')
     # 使用 cursor() 方法创建一个游标对象 cursor
     cursor = db.cursor()

     # 使用 execute()  方法执行 SQL 查询
     # cursor.execute( "select ip,full_name from devices where full_name in ('CNSHHCJX1001E','CNSHHTAF1004E','CNSHHCJX1002E','CNSHHTAF1005E','CNSHHFTX1002E','CNSHHSDS1001E','CNSHHFTX1001E','CNSHHSJG1001E')  ")
     cursor.execute( "select ip,full_name from devices ")

     # 使用 fetchone() 方法获取单条数据.
     data = cursor.fetchall()
     db.close()
     # 关闭数据库连接
     for line in data:
          print(line[0] + "," + line[1])
     return data

#根据获取的设备信息来获取ip
def getIp(data,device_name):
     for i in range(0,data.__len__()) :
          if data[i][1] == device_name:
               return data[i][0]

# 打印当前时间
def getNowTime():
    nowTime=strftime("%Y-%m-%d %H:%M:%S", localtime())
    return nowTime

if __name__ == '__main__':
    """
    read_xls 参数:excel文件名，文件下的表格名，行数，登陆用户名，密码，公钥文件
    """
    executeLog+="开始时间："+getNowTime()+"\n"
    startTime=datetime.now()
    readExcel("check.xlsx")
    pub_file = 'public.pem'
    read_xls('check.xlsx')
    print("executeLog：\n"+executeLog)
    # test()
    endTime=datetime.now()
    if deviceSize!=0:
        executeLog+="结束时间："+getNowTime()+"\n"+"查询"+str(deviceSize)+"条线路，用时"+str((endTime-startTime).seconds)+"秒,一条用时：" +str((((endTime-startTime).seconds*1.00)/deviceSize))+"秒\n"
    else:
        executeLog+="结束时间："+getNowTime()+"\n"+"查询"+str(deviceSize)+"条线路，用时"+str((endTime-startTime).seconds)+"秒\n"
    print("查询"+str(deviceSize)+"条线路，用时"+str((endTime-startTime).seconds)+"秒")
    deviceslog=open("devicesLog.log","a")
    deviceslog.write(executeLog.encode())
