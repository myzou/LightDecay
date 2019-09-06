#!/usr/bin/python
# -*- coding: UTF-8 -*-

import pymysql
import openpyxl
from time import strftime, localtime

#根据获取的设备信息来获取ip
def getIp(data,device_name):
     for i in range(0,data.__len__()) :
          if data[i][1] == device_name:
               return data[i][0]


##单个查询内容
def queryLogin():
     # 打开数据库连接
     db = pymysql.connect("218.97.9.147", "gnoc_read", "gnoc_read", "nm_shared_info", charset='utf8')
     # 使用 cursor() 方法创建一个游标对象 cursor
     cursor = db.cursor()

     # 使用 execute()  方法执行 SQL 查询
     cursor.execute(
          "select ip,full_name from devices where full_name in ('CNSHHCJX1001E','CNSHHTAF1004E','CNSHHCJX1002E','CNSHHTAF1005E','CNSHHFTX1002E','CNSHHSDS1001E','CNSHHFTX1001E','CNSHHSJG1001E')  ")

     # 使用 fetchone() 方法获取单条数据.
     data = cursor.fetchall()
     db.close()
     # 关闭数据库连接
     for line in data:
          print(line[0] + "," + line[1])
     return data

#读取execl 改变excel内容
def readExcel(filename):
     data=queryLogin();
     wb = openpyxl.load_workbook(filename)
     ws = wb.active
     for i in range(2,ws.max_row+1):
          if ws.cell(row=i, column=1).value == None:
               continue
          device_name = ws.cell(row=i, column=1).value
          ip = ws.cell(row=i, column=2).value
          port = ws.cell(row=i, column=3).value
          print(i,device_name,ip,port)
          now_ip=getIp(data,device_name)
          ws.cell(row=i,column=2,value=now_ip)
          ws.cell(row=i, column=5,value=str(getNowTime()))
     wb.save(filename)
     wb.close()

# 打印当前时间
def getNowTime():
    nowTime=strftime("%Y-%m-%d %H:%M:%S", localtime())
    return nowTime

def queryAll():
     # 打开数据库连接
     db = pymysql.connect("localhost", "root", "root", "nm_shared_info", charset='utf8')
     # 使用 cursor() 方法创建一个游标对象 cursor
     cursor = db.cursor()

     # 使用 execute()  方法执行 SQL 查询
     cursor.execute("select * from devices where full_name in ('CNSHHCJX1001E','CNSHHTAF1004E','CNSHHCJX1002E','CNSHHTAF1005E','CNSHHFTX1002E','CNSHHSDS1001E','CNSHHFTX1001E','CNSHHSJG1001E')  ")

     # 使用 fetchone() 方法获取单条数据.
     data = cursor.fetchall()
     # 关闭数据库连接
     db.close()
     print("data[1][3]  " + data[1][3])
     for line in data:
          print(line[3] + "," + line[1])
     return data


if __name__ == '__main__':
     # queryLogin();
     # queryAll()
     readExcel("check.xlsx")