#!/usr/bin/python
# -*- coding: UTF-8 -*-

import base64
import configparser
import logging
import os
import traceback

import rsa
import time
from time import strftime, localtime
import openpyxl
from openpyxl import load_workbook
from urllib.request import urlopen
import sys
import json
import random
import csv
import pymysql
from datetime import datetime
from urllib.parse import quote,unquote




##单个查询内容
def queryLogin():
     # 打开数据库连接
     db = pymysql.connect("218.97.9.147", "nss_script", "ZX4XdUpH", "nm_shared_info", charset='utf8')
     # 使用 cursor() 方法创建一个游标对象 cursor
     cursor = db.cursor()

     # 使用 execute()  方法执行 SQL 查询
     # cursor.execute( "select ip,full_name from devices where full_name in ('CNSHHCJX1001E','CNSHHTAF1004E','CNSHHCJX1002E','CNSHHTAF1005E','CNSHHFTX1002E','CNSHHSDS1001E','CNSHHFTX1001E','CNSHHSJG1001E')  ")
     cursor.execute( "select ip,full_name from devices ")

     # 使用 fetchone() 方法获取单条数据.
     data = cursor.fetchall()
     db.close()
     # 关闭数据库连接
     for line in data:
          print(line[0] + "," + line[1])
     return data

#根据获取的设备信息来获取ip
def getIp(data,device_name):
     for i in range(0,data.__len__()) :
          if data[i][1] == device_name:
               return data[i][0]


# 打印当前时间
def getNowTime():
    nowTime=strftime("%Y-%m-%d %H:%M:%S", localtime())
    return nowTime

# 根据名称读取csv文件
def read_csv_map(file_path):
    # 建立空字典
    result = []
    try:
        # 读取csv至字典
        csvFile = open(file_path, "r", encoding="UTF-8")
        reader = csv.reader(csvFile)

        for item in reader:
            result.append(item)
        csvFile.close()
        return result
    except Exception as e:
        raise


'''
logging.basicConfig(level=logging.DEBUG,#控制台打印的日志级别
                    filename='after.log',
                    filemode='a',##模式，有w和a，w就是写模式，每次都会重新写日志，覆盖之前的日志
                    #a是追加模式，默认如果不写的话，就是追加模式
                    format=
                    ''
                    #日志格式
                    )
'''

def readReult():
    try:
        data = queryLogin()
        username, password, APIURL, sign,command1, before_csv,after_csv = getConfig()
        before_result=read_csv_map(before_csv)
        if len(str(before_csv))<1:
            assert ("before_csv文件长度不足")
        before_csv=str(before_csv).replace('.csv','.log')
        if os.path.exists(before_csv):
            os.remove(before_csv)
        logging.basicConfig(format='',filemode='a', level=logging.DEBUG, filename=before_csv)
        logging.info(getNowTime())
        logging.info('=================start=================')
        for i in range(1,before_result.__len__()):
            internal_site_id=before_result[i][0]
            device = before_result[i][1]
            ce_wan_ip = before_result[i][4]
            logging.info(internal_site_id)
            tempCommand = command1.replace('<CE_WAN_IP>', ce_wan_ip)
            ip = getIp(data, device)
            logging.info(device+'>'+tempCommand)
            logging.info(str(ggwAPI(ip, tempCommand, APIURL=APIURL, username=username, password=password)).replace('\n', ''))
        logging.info('=================end=================')
    except Exception as e:
        logging.basicConfig(format='',filemode='a', level=logging.DEBUG, filename='error.log')
        print(traceback.format_exc())
        logging.error(getNowTime()+'\n'+traceback.format_exc())


#读取execl 改变excel内容
def readExcel(filename):
     data = queryLogin()
     username, password, APIURL, sign, command1,before_csv,after_csv = getConfig()
     command1 = command1.replace('CE_WAN_IP', '10.16.8.54')
     ip = getIp(data, "CNGUZYUJ1001E")
     ggwAPI(ip, command=command1, APIURL=APIURL, username=username, password=password)
     wb = openpyxl.load_workbook(filename)
     ws = wb.active
     for i in range(2,ws.max_row+1):
          if ws.cell(row=i, column=1).value == None:
               continue
          CE_WAN_IP=ws.cell(row=i, column=1).value
          device_name = ws.cell(row=i, column=1).value
          ip = ws.cell(row=i, column=2).value
          port = ws.cell(row=i, column=3).value
          print(i,device_name,ip,port)
          now_ip=getIp(data,device_name)
          ws.cell(row=i,column=2,value=now_ip)
          ws.cell(row=i, column=5,value=str(getNowTime()))
     wb.save(filename)
     wb.close()

def encode_rsa(message, APIURL,ip,command):
    publicKey="-----BEGIN RSA PUBLIC KEY-----\nMIIBCgKCAQEAh/hChlpkRA+zkkEB8ZoCrVcsYsbFSXYoTBKlrdCu0LiGeKs2T+U7\nkyZ/WZ8GP498PIucz6GYN03BWOOPe5nHWfwO05XqTid6+0ni+Bfy4Ev0FyCOQsod\nmQpH4ytgn0UOp8BZyJTwdN4rtMcuY/FFnyAsFpg9+F0DtlM/dVMj/UcWaMiZIBaa\n35XkXoPa+ng8Z7ORVOPiRHXfMGrlb9gZWF5XHN1SHNBKha1uF3wexDHVm4+k3Hvb\naZFkHrgLndaYuCPPtetnSNUOw2iaiNo7Nfze1Y4ACyfvRczHEGxFEC9X7tj/Rcy2\ngmC9JCErxDQAHitX8DJrKtoeSJN8GvZZJQIDAQAB\n-----END RSA PUBLIC KEY-----\n"
    pubkey = rsa.PublicKey.load_pkcs1(publicKey)
    #公钥从文件中读取
    # with open(pub_file, 'r') as f:
    #     print(f.read().encode())
    #     pubkey = rsa.PublicKey.load_pkcs1(f.read().encode())

    temp = str(base64.b64encode(rsa.encrypt(message.encode(), pubkey)),'utf-8')
    # 登录GGW
    login = APIURL+'/GetLoginSession/RSA?crypto_sign=' + temp
    command=command.replace(' ','%20')
    #登录PE并执行命令
    exec = APIURL + '/ExecuteCommand/RSA?crypto_sign=' + temp + \
          '&&command='+command+'&&ip={0}'.format(ip)
    return login,exec


##组装密码验证字段
def get_message(username,password,sign):
    message='username=%s&&password=%s&&sign=%s&&timestamp=%s' % (username, password, sign, str(int(time.time())))
    return message

#访问
def get_result_page(url):
    res = urlopen(url).read()
    return str(res,'utf-8')


def ggwAPI(ip,command,APIURL='http://10.180.5.135:48888',username='op1768',password='Abc1015',sign='123456'):
    mg = get_message(username, password, sign)
    login, exec = encode_rsa(mg, APIURL, ip, command)
    print("login:", login)
    print("exec:", exec)
    # print("login return:", get_result_page(login))
    # print("exec return:", get_result_page(exec))
    try:
        if str(json.loads(get_result_page(login))['code']) in ['10003', '0']:
            returnExec = get_result_page(exec)
            if str(json.loads(returnExec)['code']) in ['10003', '0']:
                returnExecStr = json.loads(returnExec)['data']
                print('data:'+returnExecStr)
                return returnExecStr
        else:
            print('登录失败,请检查账号和密码是否正确')
            return '登录失败,请检查账号和密码是否正确'
    except Exception as e:
        print(str(e))
        return '登录查询信息失败'

## 加载conf里面的配置
def getConfig(configName='config.conf'):
    config = configparser.ConfigParser()
    config.read(configName,encoding='UTF-8-sig')
    lists_header = config.sections()  # 配置组名, ['luzhuo.me', 'mysql'] # 不含'DEFAULT'
    loginOP = str(config.get('login', 'loginOP')).replace('\n','').replace(' ', '')
    loginOPPassword = str(config.get('login', 'loginOPPassword')).replace('\n','').replace(' ', '')
    before_csv = str(config.get('csv_config', 'before_csv'))
    after_csv = str(config.get('csv_config', 'after_csv'))
    ggwApiUrl = str(config.get('default_config', 'ggwApiUrl'))
    sign = str(config.get('default_config', 'sign'))
    command1 = str(config.get('default_config', 'command1'))
    return loginOP,loginOPPassword,ggwApiUrl,sign,command1,before_csv,after_csv


if __name__ == '__main__':
    # readExcel('after.xlsx')
    readReult();