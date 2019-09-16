# -*- coding: utf-8 -*-
import telnetlib
import time
from time import strftime, localtime
from datetime import datetime
from openpyxl import load_workbook
import traceback
import sys

class Logger(object):
    def __init__(self, filename='console.log', stream=sys.stdout):
        self.terminal = stream
        self.log = open(filename, 'a',encoding='utf8')
    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
    def flush(self):
        pass
sys.stdout = Logger('console.log', sys.stdout)
sys.stderr = Logger('console.log', sys.stderr)

class TelnetClient():
    def __init__(self,):
        self.tn = telnetlib.Telnet()

    # 此函数实现telnet登录主机
    def login_host(self,host_ip,username,password,second=0,contain=''):
        returnStr=""
        try:
            # self.tn = telnetlib.Telnet(host_ip,port=23)
            self.tn.open(host_ip,port=23,timeout=3)
            # 等待login出现后输入用户名，最多等待10秒
            getstr=self.tn.read_until(b'Username',timeout=3)
            Username=getstr.decode()
            if 'Username' not in Username:
                returnStr="Username无法获取"
                print(returnStr)
                return returnStr
            self.tn.write(username.encode('ascii') + b'\n')
            # 等待Password出现后输入用户名，最多等待10秒
            getstr=self.tn.read_until(b'Password',timeout=5)
            Password=getstr.decode()
            if 'Password' not in Password :
                returnStr = "Password无法获取"
                print(returnStr)
                return returnStr
            self.tn.write(password.encode('ascii') + b'\n')
            self.tn.write(b'\n')
            result=self.tn.read_until(contain, timeout=second)
            # 获取登录结果
            command_result = result.decode('ascii')
            if None==command_result or ''==command_result or ': '==command_result:
                time.sleep(3)
                self.tn.write(b'\n\n\n')
                command_result=self.tn.read_very_eager().decode('ascii')
                if 'Authentication failed' in command_result:
                    print("================", "登录失败", host_ip, "Authentication failed===============================")
                    return "登录失败" + "Authentication failed"
                if 'Request Denied' in command_result:
                    print("================", "登录失败", host_ip, "Request Denied===============================")
                    return "登录失败" + "Request Denied"
                if 'timeout' in command_result:
                    print("================", "登录失败", host_ip, "timeout==============================")
                    return "登录失败" + "timeout"
                if '>' in command_result or '#' in command_result:
                    print("================", '登录成功', host_ip, "===============================")
                    return "登录成功" + host_ip
                else:
                    print("================", '登录失败', host_ip, "===============================")
                    return "登录失败,返回内容为：（" + command_result+")"
            if '>' in command_result or '#' in command_result:
                print("================", '登录成功', host_ip, "===============================")
                return "登录成功" + host_ip
            if 'Login invalid' in command_result :
                print("================",  "登录失败", host_ip, "Login invalid===============================")
                return "登录失败"+"Login invalid"
            if 'Request Denied' in command_result:
                print("================",  "登录失败", host_ip, "Request Denied===============================")
                return "登录失败"+"Request Denied"
            if 'timeout' in command_result:
                print("================",  "登录失败", host_ip,"timeout==============================")
                return "登录失败"+"timeout"

            print("command_result："+command_result)
            return "command_result："+command_result
        except Exception as e:
            print(traceback.format_exc())
            print('网络连接失败'+host_ip)
            return '网络连接失败'


    # 此函数实现执行传过来的命令，并输出其执行结果
    def execute_some_command(self, command, second=0, contain=b''):
        try:
            # 执行命令
            if command == None or command == '':
                self.tn.write(b'\n')
            else:
                self.tn.write(command.encode('ascii') + b'\n')
            result = self.tn.read_until(contain, timeout=second)
            # 获取命令结果
            command_result = result.decode('ascii')
            print(command_result, end="")
            return command_result;
        except Exception as e:
            raise e

#根据不同ip执行对应命令
def executeByIP(host_ip,username,password,enablePwd):
    returnstr=""
    archive='archive'
    ciscoCommand = 'path ftp://CEbackup:CEpassw0rd@202.76.104.109/$h-$t\nwrite-memory\nkron policy-list ' \
                   'CE-BACKUP\ncli archive config\nkron occurrence CE-BACKUP in 1:0:0 recurring \npolicy-list ' \
                   'CE-BACKUP\nntp server 203.85.247.2\nclock timezone HKT 8\n'
    try:
        telnet_client = TelnetClient()
        returnstr=telnet_client.login_host(host_ip, username, password,second=15,contain=b'>')
        # 如果登录结果返加True，则执行命令，然后退出
        if '成功' in returnstr:
            enablePwdValidate = telnet_client.execute_some_command('enable\n'+enablePwd,second=5, contain=b'#')
            if ">" in enablePwdValidate.rsplit("\r\n", 1)[1]:
                returnstr = "\n" + host_ip + "，enable密码错误"
                print(returnstr)
            else:
                telnet_client.execute_some_command('configure\n',second=5,contain=b'(config)')
                archivetemp=telnet_client.execute_some_command(archive,second=5,contain=b'at')
                if 'Invalid input detected' in archivetemp:
                    returnstr = "无法添加配置,"+archivetemp
                    print("\n" + returnstr)
                else:
                    telnet_client.execute_some_command(ciscoCommand)
                    telnet_client.execute_some_command('exit')
                    telnet_client.tn.read_very_eager()
                    telnet_client.execute_some_command('write\n',second=20,contain=b'#')
                    showArchive = telnet_client.execute_some_command('show archive', second=20, contain=b'Most Recent')
                    showArchiveStr=showArchive.rsplit("\n", 1)[1]
                    if "Error" not in showArchive and "Most Recent" in showArchive and "ftp" in showArchiveStr:
                        returnstr = "已经完成配置"
                        print("\n"+returnstr)
                    else:
                        returnstr = "返回结果出错:" +  showArchive.rsplit(":Error", 1)[1].split("\n")[0]
                        print("\n" + returnstr)
    except Exception as e:
        returnstr = str(e)
        print("执行命令异常：" + traceback.format_exc())
    print("--------------------------------------------------------------------------------------------------------")
    return returnstr


def writer(ws,row,column,str):
    ws.cell(row=row, column=column, value=str)

def writerReturnStr(ws,row,returnStr):
    writer(ws, row, 3, getNowTime())
    writer(ws, row, 5, returnStr)



def read_xls(filename,SheetName,minrow=2,maxrow=20):
    wb = load_workbook(filename)
    ws = wb[SheetName]

    returnStr=""
    if  maxrow==20:
        maxrow=ws.max_row + 1
    ##ws.max_row +1
    for i in range (minrow,maxrow):

        if ws.cell(row=i, column=1).value == None and ws.cell(row=i, column=2).value == None:
            continue
        if ws.cell(row=i, column=5).value!=None and ('已经完成配置'==ws.cell(row=i, column=5).value or '返回结果出错'==ws.cell(row=i,column=5).value or'登录失败Request Denied' in
                ws.cell(row=i, column=5).value or'登录失败Login invalid' in ws.cell(row=i, column=5).value or
                'Authentication failed' in ws.cell(row=i, column=5).value):
            continue
        if ws.cell(row=i, column=1).value == None and ws.cell(row=i, column=2).value != None:
            returnStr="CE_WAN_IP为空"
            writerReturnStr(ws,i,returnStr)
            continue
        if ws.cell(row=i, column=1).value != None and ws.cell(row=i, column=2).value == None:
            returnStr="设备型号为空"
            writerReturnStr(ws,i,returnStr)
            continue
        if ws.cell(row=i, column=1).value != None and "cisco" not in ws.cell(row=i, column=2).value.lower():
            returnStr="设备型号不为Cisco"
            writerReturnStr(ws,i,returnStr)
            continue
        if ws.cell(row=i, column=4).value == None or ws.cell(row=i, column=4).value==3:
            writer(ws,i,4,1)
        elif ws.cell(row=i, column=5).value!='已经完成配置':
            writer(ws, i, 4, int(ws.cell(row=i, column=4).value)+1)
        print("执行第"+str(i)+"数据,ip为:"+ws.cell(row=i, column=1).value +"当前时间："+getNowTime())

        returnStr=executeByIP(ws.cell(row=i, column=1).value,username,password,enablePwd)
        writerReturnStr(ws, i, returnStr)
        wb.save(filename)
# 打印当前时间
def getNowTime():
    nowTime=strftime("%Y-%m-%d %H:%M:%S", localtime())
    return nowTime

def getMinMax(filename,SheetName):
    wb = load_workbook(filename)
    ws = wb[SheetName]
    global minrow,maxrow
    if ws.cell(row=1, column=2).value != None :
        minrow=ws.cell(row=1, column=2).value
    if ws.cell(row=2, column=2).value != None:
        maxrow=ws.cell(row=2, column=2).value


minrow=2
maxrow=2000

if __name__ == '__main__':
    startTime = datetime.now()
    print("开始：",getNowTime())
    filename = '检查.xlsx'
    SheetName = 'Sheet2'

    #test环境
    # host_ip = '10.180.5.121'
    # username = 'cpcnet'
    # password = 'cpcnet123'
    # enablePwd = 'cpcnet123'



    # getMinMax(filename,'Sheet3')

    #实际环境
    host_ip = '10.117.88.234'
    username = 'op976@ce'
    password = 'huawei852*'
    enablePwd = '123cpc'
    read_xls(filename,SheetName,minrow=minrow,maxrow=maxrow)
    # executeByIP(host_ip, username, password, enablePwd)
    endTime = datetime.now()
    print("结束：",getNowTime())

    print("用时"+str((endTime-startTime).seconds)+"秒")
